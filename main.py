from clases.partida import Partida
from clases.player import Player
from persistencia import guardar_datos, cargar_datos
from rich.console import Console
from rich.panel import Panel
from rich.style import Style
import random
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import matplotlib.pyplot as plt


console = Console()
optionStyle = Style(color="medium_violet_red", bold=True, blink=True)


def registro_player():
    nick = input("Nickname: ").strip()
    elo = input("Elo: ")
    player = Player(nick, elo)
    return player


def eliminar_player(players):
    nickname = input("Nickname: ").strip()
    nick = next((n for n in players if n.nick.lower() == nickname.lower()), None)

    if not nick:
        console.print("[bold red]Player no encontrado.[/bold red]")
        return
    confirmacion = (
        input(f"¿Estás seguro que querés eliminar a {nickname}? (s/n): ")
        .strip()
        .lower()
    )
    if confirmacion == "s":
        players.remove(nick)
        console.print(
            f"[bold green]Player '{nickname}' eliminado con éxito.[/bold green]"
        )
    else:
        console.print("[bold yellow]Operación cancelada.[/bold yellow]")


def pedir_entero(mensaje):
    while True:
        try:
            valor = int(input(mensaje))
            return valor
        except ValueError:
            console.print(
                "[bold black]Por favor, ingresá un número válido.[/bold black]"
            )


def registro_partida():
    campeon = input("Campeón usado: ")
    kill = pedir_entero("Número de Asesinatos: ")
    death = pedir_entero("Número de Muertes: ")
    assist = pedir_entero("Número de Assistencias: ")
    farm = pedir_entero("Cantidad de Minions: ")
    vision = pedir_entero("Puntuación de visión: ")
    partida = Partida(campeon, kill, death, assist, vision, farm)
    return partida


def limpiar_partida(players):
    nickname = input("Nickname: ").strip()
    nick = next((n for n in players if n.nick.lower() == nickname.lower()), None)
    if not nick:
        console.print("[bold red]Player no encontrado.[/bold red]")
        return
    nick.limpiar_partida()
    console.print(
        f"[bold green]Historial de partidas de {nickname} eliminado con éxito.[/bold green]"
    )


def agregar_partida(players):
    nickname = input("Nickname: ").strip()
    if not nickname:
        console.print("[bold red]Debés ingresar un nickname.[/bold red]")
        return
    nick = next((n for n in players if n.nick.lower() == nickname.lower()), None)

    if not nick:
        console.print("[bold red]Player no encontrado.[/bold red]")
        return
    nueva_partida = registro_partida()
    nick.agregar_partida(nueva_partida)
    console.print("[bold green]Partida registrada con éxito.[/bold green]")

def simular_partida(players):
    nickname = input("Nickname: ").strip()
    nick = next((p for p in players if p.nick.lower() == nickname.lower()), None)
    
    if not nick:
        console.print("[bold red]Player no encontrado.[/bold red]")
        return

    campeones = ["Aatrox", "Ahri", "Akali", "Akshan", "Alistar", "Ambessa", "Amumu", "Anivia", "Annie", "Aphelios", "Ashe", "Aurelion Sol", 
                 "Aurora", "Azir", "Bard", "BelVeth", "Blitzcrank", "Brand", "Braum", "Briar", "Caitlyn", "Camille", "Cassiopeia", "Cho’Gath", 
                 "Corki", "Darius", "Diana", "Dr. Mundo", "Draven", "Ekko", "Elise", "Evelynn", "Ezreal", "Fiddlesticks", "Fiora", "Fizz", "Galio", 
                 "Gangplank", "Garen", "Gnar", "Gragas", "Graves", "Gwen", "Hecarim", "Heimerdinger", "Hwei", "Illaoi", "Irelia", "Ivern", "Janna", 
                 "Jarvan IV", "Jax", "Jayce", "Jhin", "Jinx", "KSante", "KaiSa", "Kalista", "Karma", "Karthus", "Kassadin", "Katarina", "Kayle", 
                 "Kayn", "Kennen", "KhaZix", "Kindred", "Kled", "KogMaw", "LeBlanc", "Lee Sin", "Leona", "Lillia", "Lissandra", "Lucian", "Lulu", 
                 "Lux", "Malphite", "Malzahar", "Maokai", "Master Yi", "Mel", "Milio", "Miss Fortune", "Mordekaiser", "Morgana", "Naafiri", "Nami", 
                 "Nasus", "Nautilus", "Neeko", "Nidalee", "Nilah", "Nocturne", "Nunu & Willump", "Olaf", "Orianna", "Ornn", "Pantheon", "Poppy", "Pyke", 
                 "Qiyana", "Quinn", "Rakan", "Rammus", "RekSai", "Rell", "Renata Glasc", "Renekton", "Rengar", "Riven", "Rumble", "Ryze", "Samira", 
                 "Sejuani", "Senna", "Seraphine", "Sett", "Shaco", "Shen", "Shyvana", "Singed", "Sion", "Sivir", "Skarner", "Smolder", "Sona", "Soraka", 
                 "Swain", "Sylas", "Syndra", "Tahm Kench", "Taliyah", "Talon", "Taric", "Teemo", "Thresh", "Tristana", "Trundle", "Tryndamere", 
                 "Twisted Fate", "Twitch", "Udyr", "Urgot", "Varus", "Vayne", "Veigar", "VelKoz", "Vex", "Vi", "Viego", "Viktor", "Vladimir", "Volibear", 
                  "Warwick", "Wukong", "Xayah", "Xerath", "Xin Zhao", "Yasuo", "Yone", "Yorick", "Yuumi", "Yunara", "Zac", "Zed", "Zeri", "Ziggs", "Zilean", "Zoe", "Zyra", ]
    
    campeon = random.choice(campeones)
    kills = random.randint(0, 20)
    deaths = random.randint(0, 15)
    assists = random.randint(0, 20)
    farm = random.randint(50, 300)
    vision = random.randint(0, 100)

    partida = Partida(campeon, kills, deaths, assists, farm, vision)

    nick.agregar_partida(partida)
    console.print(f"[bold green]Partida simulada con éxito para {nickname} usando {campeon}.[/bold green]")

def buscar_player(players):
    nickname = input("Nickname: ").strip()
    nick = next((n for n in players if n.nick.lower() == nickname.lower()), None)
    if not nick:
        console.print("[bold red]Player no encontrado.[/bold red]")
    return nick

def exportar_historial_player(player):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Historial_{player.nick}"
    encabezados = ["campeon", "kill", "death", "assist", "farm", "vision", "fecha", "KDA"]
    ws.append(encabezados)
    for col_num, encabezados in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    for partida in player.historial_partidas:
        fila = [
            partida.fecha.strftime("%d/%m/%y"),
            partida.campeon,
            partida.kill,
            partida.death,
            partida.assist,
            partida.farm,
            partida.vision,
            round(partida.calcular_kda(), 2)
        ]
        ws.append(fila)
    
    nombre_archivo = f"{player.nick}_historial_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(nombre_archivo)
    print(f"Historial exportado exitosamente como '{nombre_archivo}'")

def graficar_kda(nick):
    if not nick.historial_partidas:
        console.print("[bold red]No hay partidas registradas para graficar.[/bold red]")
        return
    fechas = [p.fecha.strftime('%d/%m/%y') for p in nick.historial_partidas]
    kdas = [p.calcular_kda() for p in nick.historial_partidas]

    plt.figure(figsize=(10, 5)) #crea una nueva figura con un tamaño específico
    plt.plot(fechas, kdas, marker='o', linestyle='-', color='purple') #dibuja un gráfico de lineas usando las listas)
    plt.title(f"KDA por partida de {nick.nick}") #titulo del gráfico
    plt.xticks(rotation=45) #rota las etiquetas del eje x para que se vean mejor
    plt.xlabel("Fecha de la partida") #etiqueta del eje x
    plt.ylabel("KDA") #etiqueta del eje y
    plt.grid(True) #agrega una cuadrícula al gráfico
    plt.tight_layout() #ajusta el diseño para que no se superpongan los elementos
    plt.show() #muestra el gráfico


def mostrar_menu():
    opciones = [
        "Registrar un nuevo Player",
        "Agregar una partida",
        "Ver historial de partidas",
        "Calcular promedio de kda",
        "Limpiar historial de partidas",
        "Eliminar un Player",
        "Simular una partida",
        "Buscar partidas por fecha",
        "Exportar historial de partidas a Excel",
        "Graficar KDA por partida",
        "Salir del sistema"
    ]
    menu = "\n".join(
        f"[bold cyan]{i + 1}.[/bold cyan] {opcion}" for i, opcion in enumerate(opciones)
    )
    console.print(Panel(menu, title="Estadísticas de League of Legends", style="blink medium_violet_red", expand=False))
    


def main():
    players = cargar_datos("jugadores.json")

    while True:
        mostrar_menu()
        try:
            opcion = int(input("Seleccione una opción: "))
        except ValueError:
            console.print(
                "[bold black]Por favor ingrese un número válido.[/bold black]"
            )
            continue

        if opcion == 1:
            nick = registro_player()
            if nick:
                if any(p.nick == nick.nick for p in players):
                    console.print(
                        "[bold red]Ya existe un Player con ese nick.[/bold red]"
                    )
                    continue
            players.append(nick)
            console.print("[bold green]Player registrado con éxito.[/bold green]")

        elif opcion == 2:
            agregar_partida(players)

        elif opcion == 3:
            nick = buscar_player(players)
            if nick: 
                nick.resumen_historial()
                """if nick.historial_partidas:
                    for i, partida in enumerate(nick.historial_partidas, 1):
                        console.print(f"[bold magenta]{i}.[/bold magenta] {partida.resumen_partida()}")
                else:
                    console.print("[bold yellow]Este jugador no tiene partidas registradas.[/bold yellow]")
"""

        elif opcion == 4:
            nick = buscar_player(players)
            if nick:
                console.print(
                    f"[bold cyan]Promedio de KDA de {nick.nick}: {nick.promedio_kda():.2f}[/bold cyan]"
                )

        elif opcion == 5:
            limpiar_partida(players)

        elif opcion == 6:
            eliminar_player(players)

        elif opcion == 7:
            simular_partida(players)

        elif opcion == 8:
            nick = buscar_player(players)
            if nick:
                try:
                    fecha = input("Ingresá la fecha (dd/mm/yyyy): ").strip()
                    datetime.strptime(fecha, "%d/%m/%Y")  # Validar formato de fecha
                except ValueError:
                    console.print("[bold red]Formato de fecha inválido. Usá dd/mm/yyyy[/bold red]")
                    continue
                partidas = nick.partidas_por_fecha(fecha)
                if partidas:
                    for i, partida in enumerate(partidas, 1):
                        console.print(f"[bold magenta]{i}.[/bold magenta] {partida.resumen_partida()}")
                else:
                    console.print("[bold yellow]No se encontraron partidas para esa fecha.[/bold yellow]")
        elif opcion == 9:
            nick = buscar_player(players)
            if nick:
                exportar_historial_player(nick)

        elif opcion == 10:
            nick = buscar_player(players)
            if nick:
                graficar_kda(nick)

        elif opcion == 11:
            guardar_datos(players, "jugadores.json")
            console.print("Saliendo del sistema.", style=optionStyle)
            break
        else:
            console.print(
                "[bold black]Opción no válida, intente de nuevo.[/bold black]"
            )


if __name__ == "__main__":
    main()
